from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO

def get_month_name(month_num):
    months = {
        '01': 'JANUARI', '02': 'FEBRUARI', '03': 'MARET', '04': 'APRIL',
        '05': 'MEI', '06': 'JUNI', '07': 'JULI', '08': 'AGUSTUS',
        '09': 'SEPTEMBER', '10': 'OKTOBER', '11': 'NOVEMBER', '12': 'DESEMBER'
    }
    return months.get(month_num, '')

def apply_borders(cell):
    thin = Side(border_style="thin", color="000000")
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def generate_surat_masuk_excel(data, bulan, tahun, bagian_name=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "DataSuratMasuk"
    
    month_name = get_month_name(bulan)
    
    # Headers
    title_rows = [
        "PEMERINTAH KECAMATAN CIPARAY",
        "KANTOR KECAMATAN CIPARAY",
        f"BAGIAN {bagian_name if bagian_name else 'TATA USAHA'}",
        "Jl. Bojong Parigi No. 111, Desa Ciparay, Kecamatan Ciparay, Kabupaten Bandung 40381 ",
        f"DATA SURAT MASUK BULAN {month_name} TAHUN {tahun}"
    ]
    
    for i, text in enumerate(title_rows, start=2):
        cell = ws.cell(row=i, column=1)
        cell.value = text
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=14)
        cell.font = Font(name='Arial', size=14, bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Table Header
    headers = ["NO", "NO URUT", "SURAT MASUK", "TANGGAL MASUK", "KODE SURAT", "DISPOSISI"]
    # ... Simplified for brevity, need to match detailed layout?
    # The user wanted feature parity. The PHP code had complex merging.
    
    ws.merge_cells('A8:A9'); ws['A8'] = "NO"
    ws.merge_cells('B8:B9'); ws['B8'] = "NO URUT"
    ws.merge_cells('C8:F8'); ws['C8'] = "SURAT MASUK"
    ws.merge_cells('G8:G9'); ws['G8'] = "TANGGAL MASUK"
    ws.merge_cells('H8:H9'); ws['H8'] = "KODE SURAT"
    ws.merge_cells('I8:N8'); ws['I8'] = "DISPOSISI"
    
    ws['C9'] = "ALAMAT PENGIRIM"
    ws['D9'] = "NOMOR SURAT"
    ws['E9'] = "TANGGAL SURAT"
    ws['F9'] = "PERIHAL"
    
    ws['I9'] = "I"; ws['J9'] = "TGL I"
    ws['K9'] = "II"; ws['L9'] = "TGL II"
    ws['M9'] = "III"; ws['N9'] = "TGL III"
    
    for row in ws['A8:N9']:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True)
            apply_borders(cell)
            
    # Data
    row_idx = 10
    for idx, item in enumerate(data, start=1):
        ws.cell(row=row_idx, column=1, value=idx)
        ws.cell(row=row_idx, column=2, value=item.get('nomorurut_suratmasuk'))
        ws.cell(row=row_idx, column=3, value=item.get('pengirim'))
        ws.cell(row=row_idx, column=4, value=item.get('nomor_suratmasuk'))
        # ... map all fields
        ws.cell(row=row_idx, column=5, value=str(item.get('tanggalsurat_suratmasuk')))
        ws.cell(row=row_idx, column=6, value=item.get('perihal_suratmasuk'))
        ws.cell(row=row_idx, column=7, value=str(item.get('tanggalmasuk_suratmasuk')))
        ws.cell(row=row_idx, column=8, value=item.get('kode_suratmasuk'))
        ws.cell(row=row_idx, column=9, value=item.get('disposisi1'))
        ws.cell(row=row_idx, column=10, value=str(item.get('tanggal_disposisi1')))
        ws.cell(row=row_idx, column=11, value=item.get('disposisi2'))
        ws.cell(row=row_idx, column=12, value=str(item.get('tanggal_disposisi2')))
        ws.cell(row=row_idx, column=13, value=item.get('disposisi3'))
        ws.cell(row=row_idx, column=14, value=str(item.get('tanggal_disposisi3')))
        
        for col in range(1, 15):
            apply_borders(ws.cell(row=row_idx, column=col))
            
        row_idx += 1

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def generate_surat_keluar_excel(data, bulan, tahun, bagian_name=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "DataSuratKeluar"
    
    month_name = get_month_name(bulan)
    
    title_rows = [
        "PEMERINTAH KECAMATAN CIPARAY",
        "KANTOR KECAMATAN CIPARAY",
        f"BAGIAN {bagian_name if bagian_name else 'TATA USAHA'}",
        "Jl. Bojong Parigi No. 111, Desa Ciparay, Kecamatan Ciparay, Kabupaten Bandung 40381 ",
        f"DATA SURAT KELUAR BULAN {month_name} TAHUN {tahun}"
    ]
    
    for i, text in enumerate(title_rows, start=2):
        cell = ws.cell(row=i, column=1)
        cell.value = text
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
        cell.font = Font(name='Arial', size=14, bold=True)
        cell.alignment = Alignment(horizontal='center')

    headers = ["No", "NOMOR SURAT", "TANGGAL KELUAR", "KODE SURAT", "NAMA BAGIAN", "TANGGAL SURAT", "KEPADA", "PERIHAL"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=8, column=i, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        apply_borders(cell)
        
    row_idx = 9
    for idx, item in enumerate(data, start=1):
        ws.cell(row=row_idx, column=1, value=idx)
        ws.cell(row=row_idx, column=2, value=item.get('nomor_suratkeluar'))
        ws.cell(row=row_idx, column=3, value=str(item.get('tanggalkeluar_suratkeluar')))
        ws.cell(row=row_idx, column=4, value=item.get('kode_suratkeluar'))
        ws.cell(row=row_idx, column=5, value=item.get('nama_bagian'))
        ws.cell(row=row_idx, column=6, value=str(item.get('tanggalsurat_suratkeluar')))
        ws.cell(row=row_idx, column=7, value=item.get('kepada_suratkeluar'))
        ws.cell(row=row_idx, column=8, value=item.get('perihal_suratkeluar'))
        
        for col in range(1, 9):
            apply_borders(ws.cell(row=row_idx, column=col))
        row_idx += 1

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
